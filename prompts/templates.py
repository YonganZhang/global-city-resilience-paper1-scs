"""prompt_builder.py — 议会 V27.3.19 prompt 构建器(从 Master Excel 读, 单一真源)

设计(用户拍板):
- per-indicator × 10 城 batch (input ~50K tokens, 8 评委 128K context 安全)
- anchor 档位匹配格式(多对一 OK, 空档 OK)
- A 段"按平均/客观水平"(不再按最弱子维度)
- 删 country_anchors → 改两种 anchor(国家级 + 360 城市)
- R2 prompt 加坚持己见 / 调整都给理由
- Consul R2 之后终审(不是 R1.5)
- Tavily 读 _data/tavily_cache/<city>.txt 软链接资产

prompt 三个版本:
1. build_r1_prompt() — 议会 R1, 8 评委独立看
2. build_r2_prompt() — 议会 R2, 看其他 7 评委 + reasoning, 鼓励坚持己见
3. build_consul_prompt() — Consul 终审, 看 R1+R2 全部
"""
from __future__ import annotations

import os
import re
from pathlib import Path
from typing import Any

from .indicator_meta import build_role_domain_block

REPO_ROOT = Path(__file__).resolve().parents[1]

# Input metadata are intentionally not distributed with this code release.
# Callers may supply the original workbook at this relative path.
_MASTER_EXCEL_CANDIDATES = [
    REPO_ROOT / "_data" / "V27_3_19_Master_Control.xlsx",
]


def _resolve_master_excel() -> Path:
    for p in _MASTER_EXCEL_CANDIDATES:
        if p.exists():
            return p
    raise FileNotFoundError(
        f"Master Excel not found in any fallback candidate. "
        f"Please put V27_3_19_Master_Control.xlsx at: {_MASTER_EXCEL_CANDIDATES[0]}\n"
        f"Searched: {[str(p) for p in _MASTER_EXCEL_CANDIDATES]}"
    )


MASTER_EXCEL = _MASTER_EXCEL_CANDIDATES[0]  # 偏好路径 (用于 docstring / display)
TAVILY_CACHE = REPO_ROOT / "_data/tavily_cache_clean"  # V27.3.20 用户拍板: 只用百科+政府+学术段, 去 [News]

# 20 tier 定义(V27.3.21 档位匹配)
TIER_DEFS = [
    ("S+", 97.5, "Top 5% (全球最优)"),
    ("S",  92.5, "Top 10%"),
    ("A+", 87.5, "Top 15%"),
    ("A",  82.5, "Top 20%"),
    ("A-", 80.0, "Top 22%"),
    ("B+", 77.5, "Top 30%"),
    ("B",  72.5, "Top 40%"),
    ("B-", 67.5, "Top 50% (中位上)"),
    ("C+", 62.5, "Top 60%"),
    ("C",  57.5, "Median"),
    ("C-", 52.5, "Bottom 55%"),
    ("D+", 47.5, "Bottom 60%"),
    ("D",  42.5, "Bottom 50%"),
    ("D-", 37.5, "Bottom 40%"),
    ("E+", 32.5, "Bottom 30%"),
    ("E",  27.5, "Bottom 25%"),
    ("E-", 22.5, "Bottom 22%"),
    ("F+", 17.5, "Bottom 15%"),
    ("F",  12.5, "Bottom 10%"),
    ("F-",  7.5, "Bottom 7%"),
    ("BOT", 2.5, "Bottom 5% (全球最差)"),
]
TIER_ORDER_STR = " > ".join(t[0] for t in TIER_DEFS)


def load_master_excel() -> dict[str, Any]:
    """读 Master Excel Sheet 6(per-indicator prompt 真源)+ 其他 sheets.

    Fallback 链: _data/ → 老 pipeline_v25/output/ → share-public 备份.
    若全部缺失 → FileNotFoundError 含明确指引.
    """
    import openpyxl
    excel_path = _resolve_master_excel()
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # Sheet 6: 18 indicator 的 prompt 元数据
    sh6 = wb["6_Prompt_Per_Indicator_18"]
    indicators = []
    for row in sh6.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        indicators.append({
            "real_name": row[0],
            "alias_name": row[1],
            "anchor_tier_map": row[2] or "",        # V27.3.21 档位匹配
            "full_prompt_template": row[3] or "",   # generic prompt
            "city_def": row[4] or "",
            "audit_status": row[5] or "",
        })

    # Sheet 7: 议会模型
    sh7 = wb["7_Parliament_Models"]
    models = []
    for row in sh7.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        models.append({"vendor": row[0], "model": row[1], "role": row[2] if len(row) > 2 else "evaluator"})

    return {"indicators": indicators, "models": models}


def load_tavily_for_city(city: str) -> str:
    """读 _data/tavily_cache/<city>.txt 软链接, 返回原始 Tavily 输出.

    返回原始格式 (含 URL + 标题 + Summary), 给 LLM 前可调用 load_summaries_only() 精简.
    """
    city_safe = city.replace(" ", "_")
    f = TAVILY_CACHE / f"{city_safe}.txt"
    if not f.exists():
        return f"[Tavily cache missing for {city}]"
    return f.read_text(encoding="utf-8")


def load_summaries_only(tavily_raw: str) -> str:
    """从原始 Tavily 输出抽出只 Summary 字段(去 URL + 标题).

    复用 _code/migration/compress_tavily_cache.py 的逻辑,
    但 inline 这里避免跨目录 import.
    """
    out_lines = []
    in_summary = False
    summary_buf = []

    def flush():
        nonlocal summary_buf
        if summary_buf:
            joined = " ".join(s.strip() for s in summary_buf).strip()
            out_lines.append(f"Summary: {joined}")
            summary_buf = []

    for line in tavily_raw.split("\n"):
        stripped = line.strip()
        if not stripped:
            if in_summary:
                flush()
                in_summary = False
            continue
        if stripped.startswith("--- ["):
            flush()
            in_summary = False
            out_lines.append("")
            out_lines.append(stripped)
            continue
        if stripped.startswith("[") and "#" in stripped:
            flush()
            in_summary = False
            m = re.match(r"^(\[\w+#\d+\])", stripped)
            out_lines.append(m.group(1) if m else stripped)
            continue
        if stripped.startswith("URL:"):
            continue
        if stripped.startswith("Summary:"):
            in_summary = True
            summary_buf = [stripped[8:].strip()]
            continue
        if in_summary:
            summary_buf.append(stripped)
    flush()
    return "\n".join(l for l in out_lines if l).strip()


def build_tier_definitions_block() -> str:
    """渲染 20 tier 定义表 + 顺序说明(明确告诉 LLM 高低)."""
    lines = ["## 档位系统(V27.3.21, 共 20 档)", ""]
    lines.append("| 档 | mid 分 | 百分位 |")
    lines.append("|---|---|---|")
    for tier, mid, pct in TIER_DEFS:
        lines.append(f"| **{tier}** | {mid} | {pct} |")
    lines.append("")
    lines.append(f"**顺序(高 → 低)**: {TIER_ORDER_STR}")
    lines.append("")
    lines.append("⚠️ 关键规则(V27.3.21 档位匹配, **不是城市强排序**):")
    lines.append("- **多个城市可以同一档**(说明该指标该水准上有多个城市)")
    lines.append("- **某些档位可以空**(无城市达到那个水准)")
    lines.append("- 目标:把目标城市落到**最匹配**的档位(允许同档)")
    return "\n".join(lines)


def build_r1_prompt(
    indicator: dict[str, Any],
    cities: list[str],
    country_anchors_excel_row: dict | None = None,
) -> tuple[str, str]:
    """议会 R1 prompt builder.

    Args:
        indicator: Sheet 6 一行 (含 alias_name / city_def / anchor_tier_map)
        cities: 10 城 batch
        country_anchors_excel_row: 该 indicator 国家级 anchor 数据(从 Sheet 1)

    Returns:
        (system_prompt, user_prompt) 给 LLM
    """
    real_name = indicator["real_name"]
    alias = indicator["alias_name"]
    city_def = indicator["city_def"]
    anchor_tier_map = indicator["anchor_tier_map"]

    # V27.3.20 用户拍板: Tavily 改可选, 默认**不加** (避免长 prompt 让 deepseek/qwen 400).
    # 想加 Tavily 改全局开关 env LLM_INCLUDE_TAVILY=1, 或者调用方传 indicator["_include_tavily"]=True.
    include_tavily = bool(indicator.get("_include_tavily", False)) or os.getenv("LLM_INCLUDE_TAVILY", "").lower() in ("1", "true", "yes")
    tavily_blocks = []
    if include_tavily:
        # 拼 Tavily summaries (只 Summary 字段, 去 URL+标题)
        for city in cities:
            raw = load_tavily_for_city(city)
            summaries = load_summaries_only(raw)
            tavily_blocks.append(f"### {city}\n{summaries}\n")
    else:
        # 默认: 仅列城市名, LLM 用自己已有知识 + anchor 评分
        for city in cities:
            tavily_blocks.append(f"### {city}\n(无外部参考, 请用自身知识评分)\n")

    tier_block = build_tier_definitions_block()
    role_domain = build_role_domain_block(real_name)

    system = f"""{role_domain}

## 任务说明

把 {len(cities)} 个目标城市分到 21 档位之一 (允许同档, 允许空档), 针对**单一 indicator**:
**{alias}**

输入: anchor 城市清单 (仅校准, 不评) + 你自身已有知识. 外部参考资料可能为空, 不要因"没有外部证据"降分.
输出: 每个目标城市的 tier + 30-60 字中文证据 (含城市名 + 至少 1 条具体事实, 可来自你自身知识)

⚠️ 评分极性: 此 indicator 衡量**机构能力**, 档次越高 = 能力越强 (S+ 最强 / BOT 最弱).
深层 deficit/resilience 翻转由代码后处理 (门 2), 你只判断**能力**, 无需考虑底层公式方向.

## Indicator 子维度定义 (从 Master Excel Sheet 6)

{city_def}

{tier_block}

## 档位匹配 anchor 表

{anchor_tier_map}

## 你的输出格式

**严格 JSON**:
```json
{{
  "indicator": "{alias}",
  "ratings": {{
    "<City Name>": {{
      "tier": "A+",
      "evidence_strength": "strong",
      "reasoning": "30-60 字中文证据"
    }},
    ...
  }}
}}
```

**字段说明**:
- `tier`: 21 档之一 (S+ 到 BOT)
- `evidence_strength`: "strong" | "medium" | "weak"
  - **strong**: reasoning 含具体 policy / 法规 / 年份 / 真实事件 (e.g. "HK 2010 马头围塌楼")
  - **medium**: 部分证据 (含模糊数字 / 一般描述)
  - **weak**: 主要靠"感觉" / 国家级推断 / Tavily 无关
- `reasoning`: 30-60 字中文, 含城市名 + ≥1 具体事实

📌 `evidence_strength` 仅用于审计 reasoning 的证据质量；不进入数值权重。
Consul 根据完整 reasoning 判断是否剔除无证据极值。

## 规则 (专业版 7 条, 强约束)

1. **系统视角** — 把 8 子维度作整合系统评估, 不是逐项加权平均
2. **只用城市级证据** — 用目标城市自身经核实的事实 (法规 / 记录 / 真实事件), 绝不用国家级均值
3. **Anchor 夹击** — 挑 2-3 个能力夹住目标的 anchor 城市作对照, 论证目标落在哪一侧
4. **极性** — 机构能力越强 ↔ 档次越高 (deficit 指标在后处理反转, 你只判断能力)
5. **证据-或-不确定** — 每个 tier 引用具体事件 / 政策 / 法规; 证据薄弱时在 reasoning 标不确定性, 不要猜
6. **不要国家级平均** — 国家分 ≠ 城市分 (例: Manila ≠ 菲律宾均值)
7. 多城同档 / 某档空 都允许 (后处理不强制每档有城市)
8. 输出**只用 JSON**, 不要 markdown 包裹, 不要前后散文
"""

    evidence_header = "## 目标 {n} 城 + 外部参考资料 (V27.3.20: 默认不附 Tavily, LLM 用自身知识 + anchor 评分)" .format(n=len(cities))
    user = f"""{evidence_header}

{''.join(tavily_blocks)}

按 system prompt 的 anchor + city_def, 给这 {len(cities)} 城每个分到 21 档之一. 输出 JSON.
注: 没有外部 Tavily 资料**不**是降分理由 — 你应基于自身知识 + anchor 城市对照评分.
"""
    return system, user


def build_r2_prompt(
    indicator: dict[str, Any],
    cities: list[str],
    r1_packet: dict[str, dict[str, Any]],  # {model_name: {city: {tier, reasoning}}}
    country_anchors_excel_row: dict | None = None,
) -> tuple[str, str]:
    """议会 R2 prompt builder.

    R2 LLM 看 R1 全部 8 评委 + 各自 reasoning, 鼓励坚持己见或调整, 都给理由.
    """
    real_name = indicator["real_name"]
    alias = indicator["alias_name"]
    city_def = indicator["city_def"]
    anchor_tier_map = indicator["anchor_tier_map"]
    tier_block = build_tier_definitions_block()
    role_domain = build_role_domain_block(real_name)

    # 拼 R1 packet
    packet_lines = ["## 议会 R1 全部 8 评委的打分 + reasoning(已匿名为 Model_A/B/C...)"]
    packet_lines.append("")
    for city in cities:
        packet_lines.append(f"### {city}")
        for i, (model, ratings) in enumerate(sorted(r1_packet.items())):
            alias_id = chr(ord("A") + i)
            r = ratings.get(city, {})
            tier = r.get("tier", "—")
            reason = r.get("reasoning", "—")
            packet_lines.append(f"- Model_{alias_id}: {tier} — {reason}")
        packet_lines.append("")
    packet = "\n".join(packet_lines)

    system = f"""{role_domain}

V27.3.19 议会 R2(辩论轮).

## 任务

你看到全部 8 评委 R1 阶段的打分 + reasoning(匿名).
**重新评估**自己的打分:
- ✅ **可以调整**(看了其他评委 reasoning 觉得有新角度)
- ✅ **可以坚持己见**(觉得自己的 R1 是对的)
- ⚠️ **不论调整或坚持, 都必须给理由**(30-60 字中文)

## Indicator(同 R1)

{alias}: {city_def}

{tier_block}

{anchor_tier_map}

## R2 评分原则

1. **坚持己见是值得鼓励的**, 多数票不一定对
2. **不要跟风**, 也不要 anchor to mean
3. **看证据**: 哪个评委给的 reasoning 更有具体事实
4. 输出**严格 JSON**

## 输出格式

```json
{{
  "indicator": "{alias}",
  "ratings": {{
    "<City Name>": {{
      "tier": "A+",
      "decision": "adjusted" | "kept",
      "evidence_strength": "strong" | "medium" | "weak",
      "reasoning": "30-60 字: 为什么调整/坚持"
    }},
    ...
  }}
}}
```

**evidence_strength 字段说明 (审计字段, 不参与数值加权)**:
- **strong**: reasoning 含具体政策名 / 法规年份 / 真实事件 (e.g. "HK 2010 马头围塌楼" / "London 2017 Grenfell 火灾 + 2022 Building Safety Act")
- **medium**: 部分证据 (含模糊数字 / 一般描述)
- **weak**: 主要靠"感觉" / 无具体证据 / 推断
- Consul 可参考该字段审查 reasoning，但 kept 评委的 tier mid-point 始终等权平均
"""

    user = packet + f"\n\n按上面 R1 packet 重新评估, 决定坚持或调整, 都给理由 + 标 evidence_strength. 输出 JSON.\n"
    return system, user


def build_consul_prompt(
    indicator: dict[str, Any],
    cities: list[str],
    r1_packet: dict[str, dict[str, Any]],
    r2_packet: dict[str, dict[str, Any]],
) -> tuple[str, str]:
    """Consul 终审 prompt builder(R2 之后跑, 不是 R1.5).

    Consul model: gemini-2.5-pro(看 R1 + R2 全部, 给最终 tier).
    """
    real_name = indicator["real_name"]
    alias = indicator["alias_name"]
    city_def = indicator["city_def"]
    anchor_tier_map = indicator["anchor_tier_map"]
    tier_block = build_tier_definitions_block()
    role_domain = build_role_domain_block(real_name)

    # 拼 R1 + R2 packet
    def fmt_packet(packet: dict, label: str) -> str:
        lines = [f"## 议会 {label} 全部评委的打分"]
        for city in cities:
            lines.append(f"### {city}")
            for i, (model, ratings) in enumerate(sorted(packet.items())):
                alias_id = chr(ord("A") + i)
                r = ratings.get(city, {})
                tier = r.get("tier", "—")
                reason = r.get("reasoning", "—")
                decision = r.get("decision", "")
                deco = f" ({decision})" if decision else ""
                lines.append(f"- Model_{alias_id}: {tier}{deco} — {reason}")
            lines.append("")
        return "\n".join(lines)

    system = f"""{role_domain}

你是议会主席 (Consul), 看完 R1 + R2 全部 8 评委的打分 + reasoning. 🔴 **你的唯一权力是去除极值** (用户拍板 2026-05-17):

- **去除最大值评委** (1 个, 那个给最高 tier 的) — 如果它过高
- **去除最小值评委** (1 个, 那个给最低 tier 的) — 如果它过低
- **去除两端**: 同时去除最高和最低 (如果两端都有问题)
- **不去**: 8 评委一致 / 极值合理时

⚠️ **你绝对不能**: 重新打分 / 自己给一个 tier / 强加判断. 你只决定"删谁".

## Indicator

{alias}: {city_def}

{tier_block}

{anchor_tier_map}

## Consul 决策原则 (只判断"去不去极值")

1. **极值是否反常**: 最高/最低评委的 reasoning 是否站不住脚 (无证据 / 极性反 / 与多数差 ≥3 档)
2. **多数共识**: 8 评委 R2 后中间 6 个是否一致 (差 ≤2 档)
3. **看证据质量**, 不看票数. 一致但都没证据 ≠ 共识
4. **保守原则**: 不确定就 keep_all (不删)

## 输出 (严格 JSON, drop 数组只列删谁, 代码后处理算最终 tier)

🔴 **8 评委编号 1-8** (字典序: Model_A=1, B=2, C=3, D=4, E=5, F=6, G=7, H=8).
Consul 只输出"删谁的位置 + 理由".

```json
{{
  "indicator": "{alias}",
  "decisions": {{
    "<City Name>": {{
      "drop": [5],
      "reasoning": "Model_5 给 A- 但 reasoning 无具体证据反常. 30-50 字."
    }},
    "<Another City>": {{
      "drop": [],
      "reasoning": "8 评委 R2 共识 B+/A-, 极值合理 keep_all. 30-50 字."
    }},
    "<Another>": {{
      "drop": [1, 5],
      "reasoning": "Model_1 极低 + Model_5 极高都 reasoning 弱, 双侧剔除. 30-50 字."
    }}
  }}
}}
```

**字段约束**:
- `drop`: 整数数组, 元素 1-8 之间. **空数组 [] = keep_all 不删**. **最多 2 个元素** (1 max + 1 min).
  - 长度 ≥3 = 议会失败信号, 必须改 [] 让 polarity_aggregate 后处理.
- `reasoning`: 必填中文 30-50 字, 说明删谁/为什么 OR 为什么 keep_all.

**5 种 drop 典型值** (按 prompt 给的 Model_A/B/C/D/E/F/G/H 字典序对应 1-8):
| drop 值 | 含义 | 何时用 |
|---|---|---|
| `[]` | keep_all 不删 | 8 评委共识 / 极值合理 / 不确定 (默认保守) |
| `[5]` | 删 Model_E (5 号) | Model_E 极高/极低 + reasoning 无证据 |
| `[1]` | 删 Model_A (1 号) | Model_A 极低/极高 + reasoning 弱 |
| `[1, 5]` | 双删 Model_A + Model_E | 双侧反常 (极高极低都无证据) |
| `[1, 2, 3]` | 🔴 **禁止** | ≥3 是议会失败信号, 必须 [] |

**极值反常硬约束 (C.1, 双条件 AND)**:
反常 = (该评委 R2 tier 跟中位差 ≥3 档) **AND** (reasoning 无具体证据).
两条都符才算反常. 只满足一条 (e.g. 差 ≥3 档但有证据) 可能是 hyper-local 真发现 → 保留.

**后处理 (代码做, 不是 LLM 做)**: 用 drop 数组 index 删对应 Model_X, 剩余 kept 评委 R2 tier → resilience → mean → 找最近 tier 作最终输出.
"""

    user = fmt_packet(r1_packet, "R1") + "\n\n" + fmt_packet(r2_packet, "R2") + f"\n\n按上面 R1+R2 全部, 给 {alias} 的 {len(cities)} 城最终 tier. 输出 JSON.\n"
    return system, user
